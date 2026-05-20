#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для преобразования DAT файлов в формат Excel
Данные группируются по циклам (Ko), сортируются по возрастанию Трег (ॣ)
Каждый цикл начинается и заканчивается на одном и том же минимальном значении Трег
"""

import re
import pandas as pd
from openpyxl import Workbook

def parse_dat_file(filepath):
    """Извлекает данные из DAT файла"""
    data = []
    
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()
    
    # Ищем строки таблицы с данными
    # Формат: |  N|  Ko   |   To,ᥪ.    | .,ᥪ.|  ॣ, ᥪ. |   Rd   |
    pattern = r'\|\s*(\d+)\s*\|\s*([\d.]+)\s*\|\s*([\d.]+)\s*\|\s*([\d.]+)\s*\|\s*([\d.]+)\s*\|\s*([\d.]+)\s*\|'
    
    matches = re.findall(pattern, content)
    
    for match in matches:
        n = int(match[0])
        ko = float(match[1])
        to = float(match[2])
        ti = float(match[3])
        treg = float(match[4])  # ॣ - это Трег
        rd = float(match[5])
        
        data.append({
            'N': n,
            'Ko': ko,
            'To': to,
            'Ti': ti,
            'Treg': treg,
            'Rd': rd
        })
    
    return data

def process_cycle_data(data, cycle_name):
    """
    Обрабатывает данные для одного цикла:
    1. Находит минимальное значение Трег
    2. Сортирует по возрастанию Трег
    3. Формирует цикл: от минимума через все значения по возрастанию до максимума,
       затем сразу возвращается к тому же минимальному значению
    """
    if not data:
        return []
    
    # Находим минимальное значение Трег в цикле
    min_treg = min(row['Treg'] for row in data)
    
    # Находим строку с минимальным Treг (первую встреченную)
    min_row = None
    for row in data:
        if row['Treg'] == min_treg:
            min_row = row
            break
    
    # Сортируем данные по возрастанию Трег
    sorted_data = sorted(data, key=lambda x: x['Treg'])
    
    # Убираем дубликаты по Трег (оставляем первый)
    seen_treg = set()
    unique_data = []
    for row in sorted_data:
        if row['Treg'] not in seen_treg:
            seen_treg.add(row['Treg'])
            unique_data.append(row)
    
    # Формируем результат: цикл от минимума до максимума и обратно к минимуму
    result = []
    
    # Добавляем все уникальные значения по возрастанию (начиная с минимального)
    for row in unique_data:
        result.append({
            'Treg': row['Treg'],
            'Rd': row['Rd']
        })
    
    # Завершаем цикл тем же минимальным значением (с которого начали)
    result.append({
        'Treg': min_row['Treg'],
        'Rd': min_row['Rd']
    })
    
    return result

def format_number(num):
    """Форматирует число с запятой как десятичный разделитель"""
    return str(num).replace('.', ',')

def main():
    dat_files = [
        ('MOM_PI.dat', 'MOM_PI'),
        ('MOM_PID.dat', 'MOM_PID'),
        ('LAMBDA_PI.dat', 'LAMBDA_PI'),
        ('LAMBDA_PID.dat', 'LAMBDA_PID')
    ]
    
    wb = Workbook()
    wb.remove(wb.active)  # Удаляем пустой лист по умолчанию
    
    # Создаем сводный лист с диапазонами циклов
    summary_sheet = wb.create_sheet('Сводные_данные')
    summary_sheet.append(['Регулятор', 'Ko', 'Мин_Трег', 'Макс_Трег', 'Начало_цикла', 'Конец_цикла'])
    
    all_processed = {}
    
    for filepath, sheet_name in dat_files:
        try:
            data = parse_dat_file(filepath)
            
            if not data:
                print(f"Нет данных в файле {filepath}")
                continue
            
            # Группируем по Ko
            cycles = {}
            for row in data:
                ko = row['Ko']
                if ko not in cycles:
                    cycles[ko] = []
                cycles[ko].append(row)
            
            # Создаем лист для текущего регулятора
            ws = wb.create_sheet(sheet_name)
            ws.append(['Трег', 'Rd'])
            
            all_processed[sheet_name] = []
            
            # Обрабатываем каждый цикл (сортируем Ko по возрастанию)
            for ko in sorted(cycles.keys()):
                cycle_data = cycles[ko]
                processed = process_cycle_data(cycle_data, f"{sheet_name}_Ko{ko}")
                
                # Записываем в лист
                for row in processed:
                    treg_str = format_number(row['Treg'])
                    rd_str = format_number(row['Rd'])
                    ws.append([treg_str, rd_str])
                    all_processed[sheet_name].append({
                        'Ko': ko,
                        'Treg': row['Treg'],
                        'Rd': row['Rd']
                    })
                
                # Добавляем информацию в сводный лист
                if processed:
                    treg_values = [row['Treg'] for row in processed]
                    min_treg = min(treg_values)
                    max_treg = max(treg_values)
                    summary_sheet.append([
                        sheet_name,
                        format_number(ko),
                        format_number(min_treg),
                        format_number(max_treg),
                        format_number(min_treg),
                        format_number(min_treg)
                    ])
            
            print(f"Обработан файл {filepath}: {len(data)} записей, {len(cycles)} циклов")
            
        except Exception as e:
            print(f"Ошибка обработки файла {filepath}: {e}")
    
    # Сохраняем результат
    output_file = '/workspace/результаты_преобразования.xlsx'
    wb.save(output_file)
    print(f"\nРезультат сохранен в {output_file}")
    
    # Выводим пример для проверки
    print("\nПример данных (первые 15 строк MOM_PI):")
    if 'MOM_PI' in all_processed:
        for i, row in enumerate(all_processed['MOM_PI'][:15]):
            print(f"{format_number(row['Treg'])}\t{format_number(row['Rd'])}")

if __name__ == '__main__':
    main()
